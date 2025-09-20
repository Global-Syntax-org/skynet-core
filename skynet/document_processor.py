#!/usr/bin/env python3
"""
Document processing module for Skynet Core
Handles document upload, processing, and text extraction
"""

import os
import hashlib
import asyncio
import aiofiles
from typing import List, Dict, Optional, Tuple
from pathlib import Path
import logging
import mimetypes
import json
from datetime import datetime

logger = logging.getLogger(__name__)


class DocumentProcessor:
    """Handles document upload and processing"""
    
    def __init__(self, upload_dir: str = "data/documents"):
        self.upload_dir = Path(upload_dir)
        self.upload_dir.mkdir(parents=True, exist_ok=True)
        
        # Supported file types
        self.supported_types = {
            'text/plain': self._process_txt,
            'application/pdf': self._process_pdf,
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document': self._process_docx,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': self._process_xlsx,
            'text/markdown': self._process_txt,
            'text/csv': self._process_txt,
        }
        
        # Max file size (10MB default)
        self.max_file_size = 10 * 1024 * 1024
        
        # Text chunking settings
        self.chunk_size = 1000
        self.chunk_overlap = 200
    
    async def upload_document(self, file_data: bytes, filename: str, user_id: str) -> Dict:
        """Upload and process a document"""
        try:
            # Validate file size
            if len(file_data) > self.max_file_size:
                raise ValueError(f"File too large. Maximum size is {self.max_file_size // (1024*1024)}MB")
            
            # Detect MIME type
            mime_type, _ = mimetypes.guess_type(filename)
            if not mime_type or mime_type not in self.supported_types:
                raise ValueError(f"Unsupported file type: {mime_type}")
            
            # Generate unique file ID
            file_hash = hashlib.sha256(file_data).hexdigest()[:16]
            file_id = f"{user_id}_{file_hash}_{filename}"
            
            # Save file
            file_path = self.upload_dir / file_id
            async with aiofiles.open(file_path, 'wb') as f:
                await f.write(file_data)
            
            # Process document
            text_content = await self._extract_text(file_path, mime_type)
            
            # Create chunks
            chunks = self._create_chunks(text_content)
            
            # Create document metadata
            doc_metadata = {
                'id': file_id,
                'original_name': filename,
                'mime_type': mime_type,
                'size': len(file_data),
                'user_id': user_id,
                'upload_time': datetime.now().isoformat(),
                'file_path': str(file_path),
                'text_length': len(text_content),
                'chunk_count': len(chunks),
                'chunks': chunks
            }
            
            # Save metadata
            metadata_path = self.upload_dir / f"{file_id}.json"
            async with aiofiles.open(metadata_path, 'w') as f:
                await f.write(json.dumps(doc_metadata, indent=2))
            
            logger.info(f"Document processed: {filename} -> {file_id}")
            
            return {
                'success': True,
                'document_id': file_id,
                'original_name': filename,
                'text_length': len(text_content),
                'chunk_count': len(chunks),
                'message': f'Document "{filename}" uploaded and processed successfully'
            }
            
        except Exception as e:
            logger.error(f"Error processing document {filename}: {e}")
            return {
                'success': False,
                'error': str(e),
                'message': f'Failed to process document "{filename}": {str(e)}'
            }
    
    async def _extract_text(self, file_path: Path, mime_type: str) -> str:
        """Extract text from document based on MIME type"""
        processor = self.supported_types.get(mime_type)
        if not processor:
            raise ValueError(f"No processor for MIME type: {mime_type}")
        
        return await processor(file_path)
    
    async def _process_txt(self, file_path: Path) -> str:
        """Process plain text files"""
        try:
            async with aiofiles.open(file_path, 'r', encoding='utf-8') as f:
                return await f.read()
        except UnicodeDecodeError:
            # Try with different encoding
            async with aiofiles.open(file_path, 'r', encoding='latin-1') as f:
                return await f.read()
    
    async def _process_pdf(self, file_path: Path) -> str:
        """Process PDF files"""
        try:
            import PyPDF2
            
            text = ""
            with open(file_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                for page in reader.pages:
                    text += page.extract_text() + "\n"
            
            return text.strip()
        except Exception as e:
            raise RuntimeError(f"Error processing PDF: {e}")
    
    async def _process_docx(self, file_path: Path) -> str:
        """Process Word documents"""
        try:
            from docx import Document
            
            doc = Document(file_path)
            text = []
            
            for paragraph in doc.paragraphs:
                text.append(paragraph.text)
            
            return "\n".join(text)
        except Exception as e:
            raise RuntimeError(f"Error processing DOCX: {e}")
    
    async def _process_xlsx(self, file_path: Path) -> str:
        """Process Excel files"""
        try:
            import openpyxl
            
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            text = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text.append(f"Sheet: {sheet_name}\n")
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text.append(row_text)
                
                text.append("\n")
            
            return "\n".join(text)
        except Exception as e:
            raise RuntimeError(f"Error processing XLSX: {e}")
    
    def _create_chunks(self, text: str) -> List[Dict]:
        """Split text into overlapping chunks"""
        if not text.strip():
            return []
        
        chunks = []
        words = text.split()
        
        if len(words) <= self.chunk_size:
            # Text is small enough for a single chunk
            chunks.append({
                'id': 0,
                'text': text,
                'word_count': len(words),
                'start_word': 0,
                'end_word': len(words)
            })
            return chunks
        
        chunk_id = 0
        start = 0
        
        while start < len(words):
            end = min(start + self.chunk_size, len(words))
            
            chunk_text = " ".join(words[start:end])
            chunks.append({
                'id': chunk_id,
                'text': chunk_text,
                'word_count': end - start,
                'start_word': start,
                'end_word': end
            })
            
            # Move start position with overlap
            start = end - self.chunk_overlap
            if start < 0:
                start = 0
            
            chunk_id += 1
            
            # Break if we've covered all words
            if end >= len(words):
                break
        
        return chunks
    
    async def get_document_metadata(self, document_id: str) -> Optional[Dict]:
        """Get document metadata by ID"""
        try:
            metadata_path = self.upload_dir / f"{document_id}.json"
            if not metadata_path.exists():
                return None
            
            async with aiofiles.open(metadata_path, 'r') as f:
                content = await f.read()
                return json.loads(content)
        except Exception as e:
            logger.error(f"Error loading document metadata {document_id}: {e}")
            return None
    
    async def list_user_documents(self, user_id: str) -> List[Dict]:
        """List all documents for a user"""
        documents = []
        
        try:
            for metadata_file in self.upload_dir.glob(f"{user_id}_*.json"):
                try:
                    async with aiofiles.open(metadata_file, 'r') as f:
                        content = await f.read()
                        doc_data = json.loads(content)
                        
                        # Include only summary info
                        documents.append({
                            'id': doc_data['id'],
                            'original_name': doc_data['original_name'],
                            'upload_time': doc_data['upload_time'],
                            'size': doc_data['size'],
                            'text_length': doc_data['text_length'],
                            'chunk_count': doc_data['chunk_count']
                        })
                except Exception as e:
                    logger.warning(f"Error reading document metadata {metadata_file}: {e}")
                    continue
            
            # Sort by upload time (newest first)
            documents.sort(key=lambda x: x['upload_time'], reverse=True)
            
        except Exception as e:
            logger.error(f"Error listing documents for user {user_id}: {e}")
        
        return documents
    
    async def delete_document(self, document_id: str, user_id: str) -> bool:
        """Delete a document and its metadata"""
        try:
            # Verify ownership
            metadata = await self.get_document_metadata(document_id)
            if not metadata or metadata['user_id'] != user_id:
                return False
            
            # Delete files
            file_path = Path(metadata['file_path'])
            metadata_path = self.upload_dir / f"{document_id}.json"
            
            if file_path.exists():
                file_path.unlink()
            
            if metadata_path.exists():
                metadata_path.unlink()
            
            logger.info(f"Document deleted: {document_id}")
            return True
            
        except Exception as e:
            logger.error(f"Error deleting document {document_id}: {e}")
            return False
    
    async def search_documents(self, query: str, user_id: str, max_results: int = 5) -> List[Dict]:
        """Search for relevant document chunks based on query"""
        results = []
        
        try:
            user_documents = await self.list_user_documents(user_id)
            
            for doc_summary in user_documents:
                doc_metadata = await self.get_document_metadata(doc_summary['id'])
                if not doc_metadata:
                    continue
                
                # Improved text search in chunks
                for chunk in doc_metadata['chunks']:
                    chunk_text_lower = chunk['text'].lower()
                    query_lower = query.lower()
                    
                    # Split query into words for better matching
                    query_words = query_lower.split()
                    
                    # Check if any word matches or if the full query matches
                    matches = False
                    total_score = 0
                    
                    # Check for full query match
                    if query_lower in chunk_text_lower:
                        matches = True
                        total_score += chunk_text_lower.count(query_lower) * len(query_words)
                    
                    # Check for individual word matches
                    for word in query_words:
                        if word in chunk_text_lower:
                            matches = True
                            total_score += chunk_text_lower.count(word)
                    
                    if matches:
                        results.append({
                            'document_id': doc_metadata['id'],
                            'document_name': doc_metadata['original_name'],
                            'chunk_id': chunk['id'],
                            'text': chunk['text'],
                            'relevance_score': total_score
                        })
            
            # Sort by relevance score
            results.sort(key=lambda x: x['relevance_score'], reverse=True)
            
        except Exception as e:
            logger.error(f"Error searching documents for user {user_id}: {e}")
        
        return results[:10]  # Return top 10 results


# Global document processor instance
document_processor = DocumentProcessor()
